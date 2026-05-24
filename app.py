from __future__ import annotations

import random
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import streamlit as st
from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DEFAULT_SHEET_NAME = "Вопросы"

REQUIRED_COLUMNS = [
    "Номер вопроса",
    "Текст вопроса",
    "Вариант A",
    "Вариант B",
    "Вариант C",
    "Вариант D",
    "Вариант E",
    "Правильный ответ",
    "Тема",
    "Комментарий",
    "Исходный текст вопроса",
    "Статус проверки",
]

EDITABLE_COLUMNS = [
    "Текст вопроса",
    "Вариант A",
    "Вариант B",
    "Вариант C",
    "Вариант D",
    "Вариант E",
    "Правильный ответ",
    "Тема",
    "Комментарий",
    "Статус проверки",
]

OPTION_COLUMNS = ["Вариант A", "Вариант B", "Вариант C", "Вариант D", "Вариант E"]
ANSWER_MAP = {
    "A": "A",
    "B": "B",
    "C": "C",
    "D": "D",
    "E": "E",
    "А": "A",
    "Б": "B",
    "В": "C",
    "Г": "D",
    "Д": "E",
}

QuestionRow = dict[str, Any]


def ensure_directories() -> None:
    DATA_DIR.mkdir(exist_ok=True)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_answer(value: Any) -> str:
    """Приводит правильный ответ к формату A-E."""
    text = clean_text(value).upper()
    if not text:
        return ""

    return ANSWER_MAP.get(text[0], "")


def find_default_excel_files() -> list[Path]:
    files = sorted(DATA_DIR.glob("*.xlsx"))
    return [path for path in files if not path.name.startswith("~$")]


def read_excel_file(source: Any) -> tuple[list[QuestionRow], list[str]]:
    """Читает лист 'Вопросы' из Excel без pandas."""
    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as error:
        return [], [f"Не удалось открыть Excel-файл: {error}"]

    if DEFAULT_SHEET_NAME not in workbook.sheetnames:
        return [], [f"В файле нет листа '{DEFAULT_SHEET_NAME}'."]

    worksheet = workbook[DEFAULT_SHEET_NAME]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return [], ["Лист 'Вопросы' пустой."]

    headers = [clean_text(value) for value in rows[0]]
    missing_columns = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing_columns:
        return [], ["Не хватает обязательных колонок: " + ", ".join(missing_columns)]

    table: list[QuestionRow] = []
    for values in rows[1:]:
        row = {header: "" for header in REQUIRED_COLUMNS}
        raw_row = dict(zip(headers, values))
        for column in REQUIRED_COLUMNS:
            row[column] = clean_text(raw_row.get(column, ""))
        row["Правильный ответ"] = normalize_answer(row["Правильный ответ"])
        table.append(row)

    return table, []


def prepare_questions(table: list[QuestionRow]) -> tuple[list[QuestionRow], list[QuestionRow]]:
    """Отделяет вопросы, пригодные для тестирования, от проблемных."""
    ready: list[QuestionRow] = []
    problem: list[QuestionRow] = []

    for row in table:
        has_text = bool(clean_text(row.get("Текст вопроса")))
        has_options = any(clean_text(row.get(column)) for column in OPTION_COLUMNS)
        has_answer = bool(clean_text(row.get("Правильный ответ")))

        if has_text and has_options and has_answer:
            ready.append(row)
        else:
            problem.append(row)

    return ready, problem


def get_row_topic(row: QuestionRow) -> str:
    return clean_text(row.get("Тема")) or "Без темы"


def filter_by_topic(table: list[QuestionRow], topic: str) -> list[QuestionRow]:
    if topic == "Все темы":
        return table
    return [row for row in table if get_row_topic(row) == topic]


def get_options(row: QuestionRow, shuffle_options: bool = False) -> list[dict[str, str]]:
    options: list[dict[str, str]] = []

    for label, column in zip(["A", "B", "C", "D", "E"], OPTION_COLUMNS):
        text = clean_text(row.get(column))
        if text:
            options.append({"label": label, "text": text, "display": f"{label}. {text}"})

    if shuffle_options:
        random.shuffle(options)

    return options


def select_questions(
    table: list[QuestionRow],
    count: int,
    shuffle_questions: bool,
    state_key: str,
) -> list[QuestionRow]:
    count = min(count, len(table))
    signature = (tuple(clean_text(row.get("Номер вопроса")) for row in table), count, shuffle_questions)
    signature_key = f"{state_key}_signature"

    if st.session_state.get(signature_key) != signature:
        selected = list(table)
        if shuffle_questions:
            random.shuffle(selected)
        st.session_state[state_key] = selected[:count]
        st.session_state[signature_key] = signature

    return st.session_state.get(state_key, [])


def reset_mode_state(prefix: str) -> None:
    for key in [key for key in st.session_state if key.startswith(prefix)]:
        del st.session_state[key]


def result_grade(percent: float) -> str:
    if percent >= 90:
        return "отлично"
    if percent >= 75:
        return "хорошо"
    if percent >= 60:
        return "удовлетворительно"
    return "нужно повторить"


def rows_to_workbook_bytes(rows: list[QuestionRow], columns: list[str]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = DEFAULT_SHEET_NAME
    worksheet.append(columns)

    for row in rows:
        worksheet.append([row.get(column, "") for column in columns])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_result_file_name() -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"exam_result_{timestamp}.xlsx"


RESULT_COLUMNS = [
    "Дата",
    "Режим",
    "Номер вопроса",
    "Тема",
    "Текст вопроса",
    "Выбранный ответ",
    "Правильный ответ",
    "Результат",
    "Комментарий",
]


def build_result_row(row: QuestionRow, selected_answer: str, mode: str, is_correct: bool) -> QuestionRow:
    return {
        "Дата": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Режим": mode,
        "Номер вопроса": row.get("Номер вопроса", ""),
        "Тема": row.get("Тема", ""),
        "Текст вопроса": row.get("Текст вопроса", ""),
        "Выбранный ответ": selected_answer,
        "Правильный ответ": row.get("Правильный ответ", ""),
        "Результат": "Верно" if is_correct else "Ошибка",
        "Комментарий": row.get("Комментарий", ""),
    }


def show_problem_questions(problem_questions: list[QuestionRow]) -> None:
    if not problem_questions:
        return

    with st.expander("Проблемные вопросы"):
        st.warning(
            "Эти строки не попали в тестирование: нет текста, вариантов ответа или правильного ответа."
        )
        st.dataframe(problem_questions, use_container_width=True)


def show_score(prefix: str) -> None:
    solved = st.session_state.get(f"{prefix}_solved", 0)
    correct = st.session_state.get(f"{prefix}_correct", 0)
    mistakes = solved - correct
    percent = round(correct / solved * 100, 1) if solved else 0
    st.caption(f"Решено: {solved} | Верно: {correct} | Ошибок: {mistakes} | {percent}%")


def show_training_mode(questions: list[QuestionRow], shuffle_options: bool) -> None:
    st.header("Тренировка")

    if not questions:
        st.error("Нет вопросов, пригодных для тренировки.")
        return

    if "training_index" not in st.session_state:
        st.session_state.training_index = 0
        st.session_state.training_solved = 0
        st.session_state.training_correct = 0
        st.session_state.training_checked = False
        st.session_state.training_selected = ""

    index = min(st.session_state.training_index, len(questions) - 1)
    row = questions[index]
    options = get_options(row, shuffle_options)

    st.caption(f"Вопрос {index + 1} из {len(questions)}")
    show_score("training")
    st.subheader(clean_text(row.get("Текст вопроса")))

    selected_display = st.radio(
        "Выберите ответ",
        [option["display"] for option in options],
        key=f"training_radio_{index}",
    )
    selected_answer = next(
        option["label"] for option in options if option["display"] == selected_display
    )

    if st.button("Проверить", disabled=st.session_state.training_checked):
        is_correct = selected_answer == row.get("Правильный ответ")
        st.session_state.training_selected = selected_answer
        st.session_state.training_checked = True
        st.session_state.training_solved += 1
        if is_correct:
            st.session_state.training_correct += 1
        else:
            st.session_state.mistake_questions.append(row)
        st.rerun()

    if st.session_state.training_checked:
        selected_answer = st.session_state.training_selected
        if selected_answer == row.get("Правильный ответ"):
            st.success("Правильно.")
        else:
            st.error("Неправильно.")

        st.info(f"Правильный ответ: {row.get('Правильный ответ')}")
        if clean_text(row.get("Комментарий")):
            st.info(f"Комментарий: {row.get('Комментарий')}")

        if st.button("Следующий вопрос"):
            st.session_state.training_index = (index + 1) % len(questions)
            st.session_state.training_checked = False
            st.session_state.training_selected = ""
            st.rerun()


def show_exam_mode(questions: list[QuestionRow], shuffle_options: bool) -> None:
    st.header("Экзамен")

    if not questions:
        st.error("Нет вопросов, пригодных для экзамена.")
        return

    if "exam_started" not in st.session_state:
        st.session_state.exam_started = False

    if not st.session_state.exam_started:
        st.info("Нажмите кнопку ниже, чтобы начать экзамен с выбранными настройками.")
        if st.button("Начать экзамен"):
            st.session_state.exam_started = True
            st.session_state.exam_index = 0
            st.session_state.exam_answers = []
            st.session_state.exam_finished = False
            st.rerun()
        return

    if st.session_state.get("exam_finished", False):
        show_exam_result()
        return

    index = st.session_state.exam_index
    row = questions[index]
    options = get_options(row, shuffle_options)

    st.caption(f"Вопрос {index + 1} из {len(questions)}")
    st.subheader(clean_text(row.get("Текст вопроса")))

    selected_display = st.radio(
        "Выберите ответ",
        [option["display"] for option in options],
        key=f"exam_radio_{index}",
    )
    selected_answer = next(
        option["label"] for option in options if option["display"] == selected_display
    )

    button_text = "Завершить экзамен" if index + 1 == len(questions) else "Следующий вопрос"
    if st.button(button_text):
        is_correct = selected_answer == row.get("Правильный ответ")
        st.session_state.exam_answers.append(
            build_result_row(row, selected_answer, "Экзамен", is_correct)
        )

        if not is_correct:
            st.session_state.mistake_questions.append(row)

        if index + 1 == len(questions):
            st.session_state.exam_finished = True
            st.session_state.last_exam_file_name = build_result_file_name()
        else:
            st.session_state.exam_index += 1
        st.rerun()


def show_exam_result() -> None:
    rows = st.session_state.exam_answers
    total = len(rows)
    correct = sum(1 for row in rows if row.get("Результат") == "Верно")
    mistakes = total - correct
    percent = round(correct / total * 100, 1) if total else 0

    st.success("Экзамен завершён.")
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Всего вопросов", total)
    col2.metric("Правильных ответов", correct)
    col3.metric("Ошибок", mistakes)
    col4.metric("Результат", f"{percent}%")
    st.info(f"Оценка: {result_grade(percent)}")

    mistake_rows = [row for row in rows if row.get("Результат") == "Ошибка"]
    if mistake_rows:
        st.subheader("Ошибки")
        st.dataframe(
            [
                {
                    "Номер вопроса": row.get("Номер вопроса"),
                    "Текст вопроса": row.get("Текст вопроса"),
                    "Выбранный ответ": row.get("Выбранный ответ"),
                    "Правильный ответ": row.get("Правильный ответ"),
                    "Тема": row.get("Тема"),
                }
                for row in mistake_rows
            ],
            use_container_width=True,
        )
    else:
        st.success("Ошибок нет.")

    file_name = st.session_state.get("last_exam_file_name", build_result_file_name())
    st.download_button(
        "Скачать результат экзамена",
        rows_to_workbook_bytes(rows, RESULT_COLUMNS),
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.caption("Результат не сохраняется на сервере и доступен только для скачивания.")

    if st.button("Начать новый экзамен"):
        reset_mode_state("exam_")
        st.rerun()


def show_mistake_mode(shuffle_options: bool) -> None:
    st.header("Повтор ошибок")

    mistakes = st.session_state.get("mistake_questions", [])
    if not mistakes:
        st.info("Ошибок пока нет.")
        return

    unique: dict[str, QuestionRow] = {}
    for row in mistakes:
        unique[clean_text(row.get("Номер вопроса"))] = row
    questions = list(unique.values())

    if "mistake_index" not in st.session_state:
        st.session_state.mistake_index = 0
        st.session_state.mistake_checked = False
        st.session_state.mistake_selected = ""

    index = min(st.session_state.mistake_index, len(questions) - 1)
    row = questions[index]
    options = get_options(row, shuffle_options)

    st.caption(f"Ошибка {index + 1} из {len(questions)}")
    st.subheader(clean_text(row.get("Текст вопроса")))

    selected_display = st.radio(
        "Выберите ответ",
        [option["display"] for option in options],
        key=f"mistake_radio_{index}",
    )
    selected_answer = next(
        option["label"] for option in options if option["display"] == selected_display
    )

    if st.button("Проверить", disabled=st.session_state.mistake_checked):
        st.session_state.mistake_selected = selected_answer
        st.session_state.mistake_checked = True
        st.rerun()

    if st.session_state.mistake_checked:
        is_correct = st.session_state.mistake_selected == row.get("Правильный ответ")
        if is_correct:
            st.success("Правильно. Вопрос удалён из списка ошибок.")
            st.session_state.mistake_questions = [
                item
                for item in st.session_state.mistake_questions
                if clean_text(item.get("Номер вопроса")) != clean_text(row.get("Номер вопроса"))
            ]
        else:
            st.error("Пока неправильно.")
            st.info(f"Правильный ответ: {row.get('Правильный ответ')}")

        if clean_text(row.get("Комментарий")):
            st.info(f"Комментарий: {row.get('Комментарий')}")

        if st.button("Следующая ошибка"):
            st.session_state.mistake_index = 0 if is_correct else (index + 1) % len(questions)
            st.session_state.mistake_checked = False
            st.session_state.mistake_selected = ""
            st.rerun()


def normalize_editor_rows(edited_data: Any) -> list[QuestionRow]:
    if isinstance(edited_data, list):
        return [dict(row) for row in edited_data]

    if isinstance(edited_data, dict):
        row_count = max((len(values) for values in edited_data.values()), default=0)
        rows: list[QuestionRow] = []
        for index in range(row_count):
            rows.append(
                {
                    column: edited_data.get(column, [""] * row_count)[index]
                    if index < len(edited_data.get(column, []))
                    else ""
                    for column in REQUIRED_COLUMNS
                }
            )
        return rows

    return []


def show_editor_mode(table: list[QuestionRow]) -> None:
    st.header("Редактор вопросов")
    st.warning("Исходный файл на сервере не перезаписывается. Скачайте отредактированную копию.")

    edited = st.data_editor(
        table,
        column_order=REQUIRED_COLUMNS,
        disabled=[column for column in REQUIRED_COLUMNS if column not in EDITABLE_COLUMNS],
        use_container_width=True,
        num_rows="dynamic",
        height=520,
    )

    rows = normalize_editor_rows(edited) or table
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    st.download_button(
        "Скачать отредактированный Excel",
        rows_to_workbook_bytes(rows, REQUIRED_COLUMNS),
        file_name=f"edited_questions_{timestamp}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def count_by(rows: list[QuestionRow], key_func: Any) -> list[dict[str, Any]]:
    counts: dict[str, int] = {}
    for row in rows:
        key = key_func(row)
        counts[key] = counts.get(key, 0) + 1
    return [{"Значение": key, "Количество": value} for key, value in sorted(counts.items())]


def show_statistics_mode(
    table: list[QuestionRow],
    test_questions: list[QuestionRow],
    problem_questions: list[QuestionRow],
) -> None:
    st.header("Статистика")

    topics = {get_row_topic(row) for row in table}
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Всего вопросов", len(table))
    col2.metric("С правильным ответом", len(test_questions))
    col3.metric("Проблемных вопросов", len(problem_questions))
    col4.metric("Тем", len(topics))

    st.subheader("Распределение по темам")
    st.dataframe(count_by(table, get_row_topic), use_container_width=True)

    st.subheader("Статусы проверки")
    st.dataframe(
        count_by(table, lambda row: clean_text(row.get("Статус проверки")) or "Пусто"),
        use_container_width=True,
    )
    st.info("История экзаменов не хранится на сервере. Пользователь скачивает результат сразу после экзамена.")


def load_questions_ui(access_mode: str) -> tuple[list[QuestionRow], list[QuestionRow], list[QuestionRow]]:
    st.sidebar.header("Файл с вопросами")

    is_admin = access_mode == "Режим администратора"
    uploaded_file = None
    if is_admin:
        uploaded_file = st.sidebar.file_uploader(
            "Временно загрузить Excel",
            type=["xlsx"],
            help="Файл используется только в текущей сессии и не сохраняется на сервере.",
        )

    default_files = find_default_excel_files()
    selected_path: Path | None = None

    if default_files:
        file_names = [path.name for path in default_files]
        selected_name = st.sidebar.selectbox("Файл из data", file_names)
        selected_path = next(path for path in default_files if path.name == selected_name)
    else:
        st.sidebar.info("В папке data пока нет Excel-файлов.")

    source = uploaded_file if uploaded_file is not None else selected_path
    if source is None:
        st.info("Добавьте Excel-файл с вопросами в папку data.")
        return [], [], []

    table, errors = read_excel_file(source)
    if errors:
        for error in errors:
            st.error(error)
        return [], [], []

    test_questions, problem_questions = prepare_questions(table)
    return table, test_questions, problem_questions


def main() -> None:
    st.set_page_config(page_title="Тренажёр тестов", layout="wide")
    ensure_directories()

    st.title("Тренажёр тестов из Excel")

    if "mistake_questions" not in st.session_state:
        st.session_state.mistake_questions = []

    st.sidebar.header("Доступ")
    access_mode = st.sidebar.radio(
        "Режим работы",
        ["Режим пользователя", "Режим администратора"],
        horizontal=False,
    )

    table, test_questions, problem_questions = load_questions_ui(access_mode)
    if not table:
        return

    st.sidebar.header("Настройки")
    if access_mode == "Режим администратора":
        available_modes = [
            "Тренировка",
            "Экзамен",
            "Повтор ошибок",
            "Редактор вопросов",
            "Статистика",
        ]
    else:
        available_modes = ["Тренировка", "Экзамен", "Повтор ошибок"]

    mode = st.sidebar.selectbox(
        "Режим",
        available_modes,
    )

    topics = ["Все темы"] + sorted({get_row_topic(row) for row in table})
    selected_topic = st.sidebar.selectbox("Тема", topics)
    filtered_questions = filter_by_topic(test_questions, selected_topic)

    available_count = len(filtered_questions)
    default_count = min(10, available_count) if available_count else 1
    question_count = st.sidebar.number_input(
        "Количество вопросов",
        min_value=1,
        max_value=max(available_count, 1),
        value=default_count,
        step=1,
    )
    shuffle_questions = st.sidebar.checkbox("Перемешивать вопросы", value=True)
    shuffle_options = st.sidebar.checkbox("Перемешивать варианты ответов", value=False)

    selected_questions = select_questions(
        filtered_questions,
        int(question_count),
        shuffle_questions,
        f"{mode}_questions",
    )

    if access_mode == "Режим администратора":
        show_problem_questions(problem_questions)

    if mode == "Тренировка":
        show_training_mode(selected_questions, shuffle_options)
    elif mode == "Экзамен":
        show_exam_mode(selected_questions, shuffle_options)
    elif mode == "Повтор ошибок":
        show_mistake_mode(shuffle_options)
    elif mode == "Редактор вопросов" and access_mode == "Режим администратора":
        show_editor_mode(table)
    elif mode == "Статистика" and access_mode == "Режим администратора":
        show_statistics_mode(table, test_questions, problem_questions)


if __name__ == "__main__":
    main()
