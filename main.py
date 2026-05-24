from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path


BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"

COLUMNS = [
    "Номер вопроса",
    "Текст вопроса",
    "Вариант A",
    "Вариант B",
    "Вариант C",
    "Вариант D",
    "Вариант E",
    "Правильный ответ",
    "Тема",
    "Комментарий",
    "Исходный текст вопроса",
    "Статус проверки",
]

OPTION_LABELS = {
    "а": "A",
    "a": "A",
    "б": "B",
    "b": "B",
    "в": "C",
    "c": "C",
    "г": "D",
    "d": "D",
    "д": "E",
    "e": "E",
}

# Начало вопроса: "1.", "1)", "Вопрос 1", "№1".
QUESTION_START_RE = re.compile(
    r"(?im)^\s*(?:(?:вопрос)\s*)?(?:№\s*)?(\d{1,4})\s*(?:[\.\)]|\s)\s*"
)

# Структура PDF из Moodle/тестовой системы:
# "Вопрос 1" -> "Выполнен" -> "Баллов..." -> текст -> "Выберите один ответ:" -> варианты.
MOODLE_QUESTION_START_RE = re.compile(r"(?im)^\s*Вопрос\s+(\d{1,4})\s*$")
CHOICE_PROMPT_RE = re.compile(r"(?im)^\s*Выберите\s+один\s+ответ\s*:\s*$")
SERVICE_LINE_RE = re.compile(r"(?i)^\s*(?:Выполнен|Баллов\b.*)\s*$")

# Варианты ответа: "а)", "б.", "a)", "A.", и т.п.
OPTION_RE = re.compile(r"(?im)(?:^|\n)\s*([АаБбВвГгДдA-Ea-e])\s*[\).]\s+")


@dataclass
class Question:
    number: int
    text: str
    source_text: str
    options: dict[str, str] = field(default_factory=dict)
    status: str = "Не проверено"


def ensure_directories() -> None:
    """Создаёт рабочие папки, если их ещё нет."""
    INPUT_DIR.mkdir(exist_ok=True)
    OUTPUT_DIR.mkdir(exist_ok=True)


def normalize_text(text: str) -> str:
    """Приводит текст из PDF к более предсказуемому виду."""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def extract_text_from_pdf(pdf_path: Path) -> str:
    """Извлекает текст из текстового PDF через PyMuPDF."""
    try:
        import fitz  # PyMuPDF
    except ImportError as error:
        raise RuntimeError(
            "Не установлена библиотека PyMuPDF. Выполните: pip install -r requirements.txt"
        ) from error

    parts: list[str] = []

    with fitz.open(pdf_path) as document:
        for page_number, page in enumerate(document, start=1):
            # Для PDF из тестовых систем важен визуальный порядок блоков.
            # Обычный get_text("text") иногда сначала отдаёт левую колонку
            # со строками "Вопрос N", а только потом правую колонку с текстом.
            blocks = page.get_text("blocks")
            text_blocks = [block for block in blocks if block[6] == 0 and block[4].strip()]
            text_blocks.sort(key=lambda block: (round(block[1], 1), round(block[0], 1)))
            page_text = "\n".join(block[4].strip() for block in text_blocks)

            if page_text.strip():
                parts.append(page_text)
            else:
                print(f"  Страница {page_number}: текст не найден.")

    return normalize_text("\n".join(parts))


def remove_question_marker(text: str) -> str:
    """Убирает номер или слово 'Вопрос' в начале фрагмента."""
    return QUESTION_START_RE.sub("", text, count=1).strip(" \n\t.-")


def clean_multiline_text(text: str) -> str:
    """Склеивает переносы строк внутри одного текстового поля."""
    return re.sub(r"\s*\n\s*", " ", text).strip(" \n\t.-")


def remove_moodle_service_header(block: str) -> str:
    """Убирает служебные строки перед текстом вопроса."""
    lines = block.splitlines()

    if lines and MOODLE_QUESTION_START_RE.match(lines[0]):
        lines = lines[1:]

    while lines and SERVICE_LINE_RE.match(lines[0]):
        lines = lines[1:]

    return "\n".join(lines).strip()


def split_moodle_question_blocks(text: str) -> list[tuple[int | None, str, bool]]:
    """Делит PDF на вопросы по служебной строке 'Вопрос N'."""
    matches = list(MOODLE_QUESTION_START_RE.finditer(text))
    if not matches:
        return []

    blocks: list[tuple[int | None, str, bool]] = []
    for index, match in enumerate(matches):
        start = match.start()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        block = text[start:end].strip()
        if block:
            blocks.append((int(match.group(1)), block, True))

    return blocks


def split_into_question_blocks(text: str) -> list[tuple[int | None, str, bool]]:
    """
    Делит общий текст на фрагменты вопросов.

    Возвращает номер из PDF, исходный фрагмент и признак уверенного разбиения.
    Если явных номеров нет, пробует разделить текст по пустым строкам.
    """
    moodle_blocks = split_moodle_question_blocks(text)
    if moodle_blocks:
        return moodle_blocks

    matches = list(QUESTION_START_RE.finditer(text))

    if matches:
        blocks: list[tuple[int | None, str, bool]] = []
        for index, match in enumerate(matches):
            start = match.start()
            end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
            block = text[start:end].strip()
            if block:
                blocks.append((int(match.group(1)), block, True))
        return blocks

    # Запасной вариант: не выбрасываем текст, а сохраняем для ручной проверки.
    fallback_blocks = [part.strip() for part in re.split(r"\n\s*\n", text) if part.strip()]
    if fallback_blocks:
        return [(None, block, False) for block in fallback_blocks]

    return []


def parse_labeled_options(options_text: str) -> tuple[dict[str, str], bool]:
    """Разбирает варианты с явными буквенными метками."""
    matches = list(OPTION_RE.finditer(options_text))

    if not matches:
        return {}, False

    options: dict[str, str] = {}
    repeated_or_unknown_label = False

    for index, match in enumerate(matches):
        raw_label = match.group(1).lower()
        label = OPTION_LABELS.get(raw_label)
        option_start = match.end()
        option_end = (
            matches[index + 1].start() if index + 1 < len(matches) else len(options_text)
        )
        option_text = options_text[option_start:option_end].strip(" \n\t.-")

        if not label or label in options:
            repeated_or_unknown_label = True
            continue

        options[label] = clean_multiline_text(option_text)

    confident = len(options) >= 2 and not repeated_or_unknown_label
    return options, confident


def parse_unlabeled_options(options_text: str) -> tuple[dict[str, str], bool]:
    """
    Разбирает варианты без буквенных меток.

    В таких PDF каждый вариант часто расположен отдельной строкой. Это менее надёжно,
    потому что длинный вариант может переноситься на несколько строк, поэтому статус
    такого вопроса позже будет "Проверить вручную".
    """
    option_lines = [
        line.strip(" \t.-")
        for line in options_text.splitlines()
        if line.strip(" \t.-")
    ]

    options: dict[str, str] = {}
    for label, option_text in zip(["A", "B", "C", "D", "E"], option_lines):
        options[label] = option_text

    return options, False


def parse_moodle_options(block: str) -> tuple[str, dict[str, str], bool]:
    """Отделяет вопрос и варианты в PDF со строкой 'Выберите один ответ:'."""
    cleaned_block = remove_moodle_service_header(block)
    choice_match = CHOICE_PROMPT_RE.search(cleaned_block)

    if not choice_match:
        return clean_multiline_text(cleaned_block), {}, False

    question_text = clean_multiline_text(cleaned_block[: choice_match.start()])
    options_text = cleaned_block[choice_match.end() :].strip()

    options, confident = parse_labeled_options(options_text)
    if not options:
        options, confident = parse_unlabeled_options(options_text)

    confident = bool(question_text) and confident
    return question_text, options, confident


def parse_options(block: str) -> tuple[str, dict[str, str], bool]:
    """Отделяет текст вопроса от вариантов ответа."""
    if MOODLE_QUESTION_START_RE.match(block.splitlines()[0].strip() if block.splitlines() else ""):
        return parse_moodle_options(block)

    options, confident_options = parse_labeled_options(block)

    if not options:
        return remove_question_marker(block), {}, False

    first_option_match = OPTION_RE.search(block)
    question_text = remove_question_marker(block[: first_option_match.start()])
    confident = bool(question_text) and confident_options
    return question_text, options, confident


def parse_questions(text: str) -> list[Question]:
    """Преобразует текст PDF в список вопросов для Excel."""
    questions: list[Question] = []

    for fallback_number, (pdf_number, block, confident_split) in enumerate(
        split_into_question_blocks(text), start=1
    ):
        question_text, options, confident_options = parse_options(block)
        question_number = pdf_number if pdf_number is not None else fallback_number

        status = "Не проверено"
        if not confident_split or not confident_options:
            status = "Проверить вручную"

        questions.append(
            Question(
                number=question_number,
                text=question_text,
                options=options,
                source_text=block,
                status=status,
            )
        )

    return questions


def save_questions_to_excel(questions: list[Question], output_path: Path) -> None:
    """Создаёт Excel-файл с вопросами и удобным оформлением."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as error:
        raise RuntimeError(
            "Не установлена библиотека openpyxl. Выполните: pip install -r requirements.txt"
        ) from error

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Вопросы"

    worksheet.append(COLUMNS)

    for question in questions:
        worksheet.append(
            [
                question.number,
                question.text,
                question.options.get("A", ""),
                question.options.get("B", ""),
                question.options.get("C", ""),
                question.options.get("D", ""),
                question.options.get("E", ""),
                "",
                "",
                "",
                question.source_text,
                question.status,
            ]
        )

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "A": 16,
        "B": 45,
        "C": 32,
        "D": 32,
        "E": 32,
        "F": 32,
        "G": 32,
        "H": 18,
        "I": 22,
        "J": 28,
        "K": 60,
        "L": 20,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for row_number in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_number].height = 42 if row_number > 1 else 30

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    workbook.save(output_path)


def process_pdf(pdf_path: Path) -> None:
    """Обрабатывает один PDF и сохраняет отдельный Excel-файл."""
    print(f"Обрабатывается файл: {pdf_path.name}")

    try:
        text = extract_text_from_pdf(pdf_path)

        if not text:
            print(
                "  Текст не извлечён. Вероятно, файл является сканом "
                "и требует OCR-распознавания."
            )
            return

        questions = parse_questions(text)

        if not questions:
            print("  Вопросы не найдены. Проверьте структуру текста в PDF.")
            return

        output_path = OUTPUT_DIR / f"{pdf_path.stem}.xlsx"
        try:
            save_questions_to_excel(questions, output_path)
        except PermissionError:
            print(
                f"  Не удалось сохранить файл: {output_path}\n"
                "  Закройте этот Excel-файл, если он открыт, и запустите программу ещё раз."
            )
            return

        print(f"  Найдено вопросов: {len(questions)}")
        print(f"  Результат сохранён: {output_path}")
    except Exception as error:
        print(f"  Ошибка при обработке файла {pdf_path.name}: {error}")


def main() -> None:
    ensure_directories()

    pdf_files = sorted(INPUT_DIR.glob("*.pdf"))
    if not pdf_files:
        print(f"PDF-файлы не найдены. Положите файлы в папку: {INPUT_DIR}")
        return

    for pdf_path in pdf_files:
        process_pdf(pdf_path)


if __name__ == "__main__":
    main()
